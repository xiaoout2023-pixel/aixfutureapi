import httpx
import json
import re
import logging
import time
import io
from typing import List, Dict, Any, Optional
from crawler.base import BaseCrawler, normalize_provider_name

logger = logging.getLogger(__name__)

ARENA_BOARD_NAMES = {
    "T2VBoard": "multimodal_t2v",
    "I2VBoard": "multimodal_i2v",
    "IEBoard": "multimodal_edit",
    "R2VBoard": "multimodal_r2v",
    "T2IBoard": "multimodal_image",
    "TTSBoard": "multimodal_tts",
}

EXCEL_SHEET_MAPPING = {
    "总排行榜": "general_overall",
    "推理模型总排行榜": "general_reasoning",
    "基础模型总排行榜": "general_base",
    "推理任务总排行榜": "general_reasoning_task",
    "开源排行榜": "general_opensource",
    "小模型10B榜": "general_small_10b",
    "小模型5B榜": "general_small_5b",
}

EXCEL_DATA_PATHS = {
    "general": "/data/generalboard/",
    "multimodal_vlm": "/data/multimodal_list/VLM/",
    "multimodal_image": "/data/multimodal_list/Image/",
    "multimodal_t2v": "/data/multimodal_list/T2V/",
    "multimodal_i2v": "/data/multimodal_list/I2V/",
    "multimodal_edit": "/data/multimodal_list/Edit/",
    "multimodal_r2v": "/data/multimodal_list/R2V/",
    "multimodal_comicshorts": "/data/multimodal_list/comic/",
    "multimodal_world": "/data/multimodal_list/World/",
    "multimodal_voice_av": "/data/multimodal_list/VoiceAV/",
    "multimodal_voice_chat": "/data/multimodal_list/VoiceChat/",
    "multimodal_tts": "/data/multimodal_list/TTS/",
    "multimodal_v": "/data/multimodal_list/V/",
    "multimodal_vlr": "/data/multimodal_list/VLR/",
}

AVAILABLE_DATES = [
    "2026年3月", "2025年度测评", "2025年11月", "2025年9月",
    "2025年7月", "2025年5月", "2025年3月", "2024年12月",
    "2024年10月", "2024年8月", "2024年6月", "2024年4月",
    "2024年2月", "2023年12月", "2023年11月", "2023年10月", "2023年9月",
]


class SuperCLUECrawler(BaseCrawler):
    def __init__(self):
        super().__init__(provider="superclue", base_url="https://www.superclueai.com")

    async def _fetch(self, url: str, binary: bool = False) -> Any:
        for attempt in range(self.max_retries):
            try:
                async with httpx.AsyncClient(timeout=60, follow_redirects=True) as client:
                    response = await client.get(url, headers=self.headers)
                    response.raise_for_status()
                    if binary:
                        logger.info(f"Fetched {url}, length: {len(response.content)} bytes (binary)")
                        return response.content
                    logger.info(f"Fetched {url}, length: {len(response.text)}")
                    return response.text
            except Exception as e:
                logger.warning(f"Attempt {attempt + 1} failed for {url}: {e}")
                if attempt == self.max_retries - 1:
                    raise
                time.sleep(2 ** attempt)
        return b"" if binary else ""

    async def _crawl_arena_boards(self) -> List[Dict]:
        logger.info("Crawling arena boards from JS bundle...")
        html = await self._fetch(self.base_url)
        js_files = re.findall(r'(?:src|href)="(/assets/[^"]+\.js)"', html)

        all_entries = []
        for js_file in js_files:
            full_url = self.base_url + js_file
            try:
                js_text = await self._fetch(full_url)
                if any(k in js_text for k in ["T2VBoard", "IEBoard", "rows:[{rank"]):
                    logger.info(f"  Found arena data in {js_file}")
                    entries = self._extract_board_data(js_text)
                    all_entries.extend(entries)
            except Exception as e:
                logger.warning(f"  Error processing {js_file}: {e}")

        return all_entries

    def _extract_board_data(self, js_text: str) -> List[Dict]:
        results = []
        for board_name, category_key in ARENA_BOARD_NAMES.items():
            pattern = rf'{board_name}\s*[,=].*?data\s*\(\s*\)\s*\{{\s*return\s*\{{\s*rows:\s*\['
            match = re.search(pattern, js_text, re.DOTALL)
            if not match:
                pattern = rf'name:\s*["\']?{board_name}["\']?\s*,\s*data\s*\(\s*\)\s*\{{\s*return\s*\{{\s*rows:\s*\['
                match = re.search(pattern, js_text, re.DOTALL)

            if not match:
                logger.warning(f"Board {board_name} not found, skipping")
                continue

            start = match.end() - 1
            bracket_count = 0
            end = start
            for i in range(start, min(start + 50000, len(js_text))):
                if js_text[i] == '[':
                    bracket_count += 1
                elif js_text[i] == ']':
                    bracket_count -= 1
                    if bracket_count == 0:
                        end = i + 1
                        break

            array_text = js_text[start:end]
            try:
                rows = json.loads(array_text)
            except json.JSONDecodeError:
                rows = self._parse_js_array(array_text)

            for row in rows:
                entry = self._normalize_entry(row, category_key)
                if entry:
                    results.append(entry)

            logger.info(f"  Board {board_name} ({category_key}): {len(rows)} entries")
        return results

    def _parse_js_array(self, text: str) -> List[Dict]:
        results = []
        obj_pattern = re.findall(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}', text)
        for obj_text in obj_pattern:
            try:
                obj = {}
                for kv in re.findall(r'(\w+)\s*:\s*([^,}\s]+|"[^"]*"|\'[^\']*\'|\[[^\]]*\])', obj_text):
                    key, value = kv
                    value = value.strip('"').strip("'")
                    try:
                        obj[key] = json.loads(value)
                    except:
                        obj[key] = value
                if obj:
                    results.append(obj)
            except:
                pass
        return results

    async def _crawl_excel_boards(self) -> List[Dict]:
        logger.info("Crawling Excel-based boards...")
        all_entries = []
        date = AVAILABLE_DATES[0]
        logger.info(f"Using date: {date}")

        for board_key, data_path in EXCEL_DATA_PATHS.items():
            xlsx_url = self.base_url + data_path + date + ".xlsx"
            try:
                content = await self._fetch(xlsx_url, binary=True)
                if not content or len(content) < 100:
                    logger.warning(f"  {board_key}: empty or too small response")
                    continue

                entries = self._parse_excel(content, board_key, date)
                all_entries.extend(entries)
                logger.info(f"  {board_key}: {len(entries)} entries from Excel")

            except Exception as e:
                logger.warning(f"  {board_key}: error - {e}")

        return all_entries

    def _parse_excel(self, content: bytes, category: str, date: str) -> List[Dict]:
        try:
            import openpyxl
        except ImportError:
            logger.error("openpyxl not installed, cannot parse Excel. Run: pip install openpyxl")
            return []

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        results = []

        for sheet_name in wb.sheetnames:
            sheet_category = EXCEL_SHEET_MAPPING.get(sheet_name, category)
            ws = wb[sheet_name]

            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue

            headers = [str(h).strip() if h else "" for h in rows[0]]

            for row_data in rows[1:]:
                if not row_data or not row_data[0]:
                    continue

                row_dict = {}
                for i, val in enumerate(row_data):
                    if i < len(headers) and headers[i]:
                        row_dict[headers[i]] = val

                entry = self._normalize_excel_entry(row_dict, sheet_category, date)
                if entry:
                    results.append(entry)

        wb.close()
        return results

    def _normalize_excel_entry(self, row: Dict, category: str, date: str) -> Optional[Dict]:
        rank = self._safe_int(row.get("排名", row.get("Rank", 0)))
        model_name = str(row.get("模型名称", row.get("模型", row.get("Model", "")))).strip()
        org = normalize_provider_name(str(row.get("机构", row.get("Organization", row.get("Org", "")))).strip())

        score = 0.0
        for score_key in ["总分", "总 分", "综合得分", "Score", "score"]:
            if score_key in row and row[score_key] is not None:
                score = self._safe_float(row[score_key])
                break

        score_details = {}
        skip_keys = {"排名", "Rank", "模型名称", "模型", "Model", "机构", "Organization", "Org",
                      "发布日期", "Date", "开源/闭源", "License", "使用方式", "Usage", "属地",
                      "总分", "总 分", "综合得分", "Score", "score", "是否推理"}
        for key, value in row.items():
            if key in skip_keys or value is None:
                continue
            if isinstance(value, (int, float)):
                score_details[key] = value
            elif isinstance(value, str):
                try:
                    score_details[key] = float(value)
                except ValueError:
                    score_details[key] = value

        open_source = str(row.get("开源/闭源", row.get("License", "")))
        is_opensource = 1 if "开源" in open_source else 0

        usage = str(row.get("使用方式", row.get("Usage", "API"))).lower()
        usage_type = "api"
        if "网页" in usage or "web" in usage:
            usage_type = "web"
        elif "模型" in usage or "开源" in usage:
            usage_type = "model"

        is_reasoning = 0
        if category in ("general_reasoning", "general_reasoning_task"):
            is_reasoning = 1
        reasoning_val = row.get("是否推理", "")
        if reasoning_val and str(reasoning_val).strip() in ("是", "Yes", "yes", "1", "True", "true"):
            is_reasoning = 1
        model_lower = model_name.lower()
        if any(k in model_lower for k in ["reasoning", "thinking", "r1", "o1", "o3", "o4", "qwq", "zero"]):
            is_reasoning = 1

        release_date = str(row.get("发布日期", row.get("Date", ""))).strip()

        if not model_name or model_name == "None":
            return None

        return {
            "category": category,
            "rank": rank,
            "model_name": model_name,
            "organization": org,
            "score": score,
            "score_details": json.dumps(score_details, ensure_ascii=False),
            "is_opensource": is_opensource,
            "is_domestic": 1 if self._is_domestic_org(org) else 0,
            "release_date": release_date,
            "usage_type": usage_type,
            "is_reasoning": is_reasoning,
        }

    def _normalize_entry(self, row: Dict, category: str) -> Dict:
        rank = self._safe_int(row.get("rank", row.get("排名", 0)))
        model_name = row.get("model", row.get("模型名称", row.get("model_name", "")))
        org = normalize_provider_name(row.get("org", row.get("organization", row.get("机构", ""))))
        date = str(row.get("date", row.get("发布日期", row.get("release_date", ""))))

        score = 0.0
        if "median" in row:
            score = self._safe_float(row["median"])
        elif "score" in row:
            score = self._safe_float(row["score"])
        elif "总分" in row:
            score = self._safe_float(row["总分"])

        score_details = {}
        for key, value in row.items():
            if key not in ("rank", "model", "org", "date", "median", "ciLow", "ciHigh", "battles"):
                if isinstance(value, (int, float)) and key not in ("rank",):
                    score_details[key] = value
                elif isinstance(value, str) and key not in ("model", "org", "date"):
                    try:
                        score_details[key] = float(value)
                    except:
                        score_details[key] = value

        if "ciLow" in row and "ciHigh" in row:
            score_details["ci_low"] = self._safe_float(row["ciLow"])
            score_details["ci_high"] = self._safe_float(row["ciHigh"])
        if "battles" in row:
            score_details["battles"] = row["battles"]

        open_source = str(row.get("license", row.get("开源/闭源", row.get("is_opensource", ""))))
        is_opensource = 1 if "开源" in open_source or open_source.lower() == "open" else 0

        usage = str(row.get("usage", row.get("使用方式", row.get("使用", "api")))).lower()
        usage_type = "api"
        if "网页" in usage or "web" in usage:
            usage_type = "web"
        elif "模型" in usage or "model" in usage or "开源" in usage:
            usage_type = "model"
        elif "api" in usage:
            usage_type = "api"

        is_reasoning = 0
        if category in ("general_reasoning", "general_reasoning_task"):
            is_reasoning = 1
        model_lower = model_name.lower()
        if any(k in model_lower for k in ["reasoning", "thinking", "r1", "o1", "o3", "o4", "qwq", "zero"]):
            is_reasoning = 1

        if not model_name:
            return None

        return {
            "category": category,
            "rank": rank,
            "model_name": model_name,
            "organization": org,
            "score": score,
            "score_details": json.dumps(score_details, ensure_ascii=False),
            "is_opensource": is_opensource,
            "is_domestic": 1 if self._is_domestic_org(org) else 0,
            "release_date": date,
            "usage_type": usage_type,
            "is_reasoning": is_reasoning,
        }

    async def crawl(self) -> List[Dict[str, Any]]:
        logger.info("Starting SuperCLUE leaderboard crawl...")

        arena_entries = await self._crawl_arena_boards()
        logger.info(f"Arena boards: {len(arena_entries)} entries")

        excel_entries = await self._crawl_excel_boards()
        logger.info(f"Excel boards: {len(excel_entries)} entries")

        all_entries = arena_entries + excel_entries
        logger.info(f"Total leaderboard entries collected: {len(all_entries)}")
        return all_entries

    @staticmethod
    def _safe_int(val) -> int:
        try:
            return int(float(str(val).replace(",", "")))
        except (ValueError, TypeError):
            return 0

    @staticmethod
    def _safe_float(val) -> float:
        try:
            return float(str(val).replace(",", "").replace("+", "").split("-")[0].strip())
        except (ValueError, TypeError):
            return 0.0

    @staticmethod
    def _is_domestic_org(org: str) -> bool:
        domestic_keywords = [
            "字节跳动", "百度", "阿里", "腾讯", "智谱", "商汤", "快手", "讯飞",
            "深度求索", "DeepSeek", "零一万物", "百川", "MiniMax", "月之暗面",
            "生数", "潞晨", "爱诗", "小米", "华为", "Baidu", "Alibaba",
            "Tencent", "Zhipu", "SenseTime", "ByteDance", "Kuaishou",
            "iFlytek", "01.AI", "Baichuan", "Moonshot", "昆仑万维"
        ]
        org_lower = org.lower()
        for kw in domestic_keywords:
            if kw.lower() in org_lower:
                return True
        return False
